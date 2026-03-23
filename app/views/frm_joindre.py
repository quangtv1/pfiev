"""Excel file merger utility — like VB frmJoindre.
Merges multiple .xlsx files (same structure) into one output file.
Header taken from first file; subsequent files' headers are skipped.
"""
from pathlib import Path
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QListWidget, QPushButton,
    QProgressBar, QFileDialog, QMessageBox, QLabel
)
from app.localization import L


class FrmJoindre(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(L.get("TitleJoindre"))
        self.setFixedSize(520, 380)
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)

        layout.addWidget(QLabel(L.get("LblFichiersAJoindre")))

        # File list + side buttons
        list_row = QHBoxLayout()
        self._lst = QListWidget()
        self._lst.currentRowChanged.connect(self._update_buttons)
        list_row.addWidget(self._lst)

        side = QVBoxLayout()
        self._btn_add = QPushButton(L.get("BtnAjouter"))
        self._btn_add.clicked.connect(self._on_add)
        self._btn_remove = QPushButton(L.get("BtnEnlever"))
        self._btn_remove.setEnabled(False)
        self._btn_remove.clicked.connect(self._on_remove)
        side.addWidget(self._btn_add)
        side.addWidget(self._btn_remove)
        side.addStretch()
        list_row.addLayout(side)
        layout.addLayout(list_row)

        # Progress bar (hidden until merge starts)
        self._progress = QProgressBar()
        self._progress.setVisible(False)
        layout.addWidget(self._progress)

        # Bottom buttons
        bottom = QHBoxLayout()
        self._btn_merge = QPushButton(L.get("BtnRealiser"))
        self._btn_merge.setEnabled(False)
        self._btn_merge.clicked.connect(self._on_merge)
        btn_cancel = QPushButton(L.get("BtnCancel"))
        btn_cancel.clicked.connect(self.reject)
        bottom.addStretch()
        bottom.addWidget(self._btn_merge)
        bottom.addWidget(btn_cancel)
        layout.addLayout(bottom)

    def _update_buttons(self):
        has_selection = self._lst.currentRow() >= 0
        count = self._lst.count()
        self._btn_remove.setEnabled(has_selection)
        self._btn_merge.setEnabled(count >= 2)

    def _on_add(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self, L.get("TitleSelectExcelFile"), "", "Excel Files (*.xlsx)"
        )
        for path in paths:
            # Skip duplicates
            existing = [self._lst.item(i).text() for i in range(self._lst.count())]
            if path not in existing:
                self._lst.addItem(path)
        self._update_buttons()

    def _on_remove(self):
        row = self._lst.currentRow()
        if row >= 0:
            self._lst.takeItem(row)
        self._update_buttons()

    def _on_merge(self):
        from datetime import date
        today = date.today().strftime("%Y%m%d")
        save_path, _ = QFileDialog.getSaveFileName(
            self, L.get("MsgJoindreSave"),
            f"merged_{today}.xlsx", "Excel Files (*.xlsx)"
        )
        if not save_path:
            return

        files = [self._lst.item(i).text() for i in range(self._lst.count())]
        try:
            from openpyxl import load_workbook, Workbook
            wb_out = Workbook()
            ws_out = wb_out.active
            ws_out.title = "Merged"

            total_files = len(files)
            self._progress.setRange(0, total_files)
            self._progress.setValue(0)
            self._progress.setVisible(True)

            header_written = False
            for idx, file_path in enumerate(files):
                if not Path(file_path).exists():
                    QMessageBox.warning(self, "", f"File not found:\n{file_path}")
                    continue
                wb_in = load_workbook(file_path, read_only=True, data_only=True)
                ws_in = wb_in.active
                first_data_row = 1  # skip header on subsequent files
                for row_idx, row in enumerate(ws_in.iter_rows(values_only=True), 1):
                    if row_idx == 1:
                        if not header_written:
                            ws_out.append(list(row))
                            header_written = True
                        # always skip source header on subsequent files
                        continue
                    ws_out.append(list(row))
                wb_in.close()
                self._progress.setValue(idx + 1)

            wb_out.save(save_path)
            self._progress.setVisible(False)
            QMessageBox.information(
                self, "",
                L.fmt("MsgJoindreResult", total_files)
            )
            self.accept()
        except Exception as e:
            self._progress.setVisible(False)
            QMessageBox.critical(self, "", str(e))
