from dataclasses import dataclass, field
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


class ExcelService:

    @staticmethod
    def export_candidats(rows: list, file_path: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "Thi sinh"

        headers = ["ID", "Họ tên", "Tên", "Ngày sinh",
                   "Giới tính", "Trạng thái", "Ngôn ngữ", "Trường"]
        keys = ["CandidatID", "Nom", "Prenom", "DateDeNaissance",
                "Sexe", "CandidatStatut", "Langue", "EtabNom"]

        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="ADD8E6")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for r, row in enumerate(rows, 2):
            for c, key in enumerate(keys, 1):
                ws.cell(row=r, column=c, value=str(row.get(key) or ""))

        # Set approximate column widths
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

        wb.save(file_path)

    @staticmethod
    def export_resultats(rows: list, file_path: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "Ket qua"
        ws["A1"] = "Kết quả kỳ thi PFIEV"
        ws["A1"].font = Font(bold=True, size=14)

        headers = ["Xếp hạng", "Họ tên", "Tên", "Trạng thái",
                   "Điểm TB", "Ngành", "Trường"]
        keys = ["CandidatClassement", "Nom", "Prenom",
                "CandidatStatut", "CandidatMoyenne", "FiliereNom", "EtabFiliere"]

        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="FFFFE0")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for r, row in enumerate(rows, 4):
            for c, key in enumerate(keys, 1):
                ws.cell(row=r, column=c, value=row.get(key))

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

        wb.save(file_path)

    @dataclass
    class ImportRow:
        nom: str = ""
        nom_intermediaire: str = ""
        prenom: str = ""
        date_de_naissance: str = ""
        sexe: str = ""
        statut: str = ""
        langue: str = ""
        etab_code: str = ""
        # filiere_choices: list of (filiere_code, priority_rank) sorted by rank asc
        filiere_choices: list = field(default_factory=list)

    @staticmethod
    def import_candidats(file_path: str) -> list:
        """Import candidates from Excel.

        Expected column order (1-indexed):
          1=No, 2=CandidatNom, 3=CandidatNomIntermediaire, 4=CandidatPrenom,
          5=CandidatDateNaissance, 6=CandidatSexe, 7=CandidatLangue,
          8=CandidatStatut, 9=EtabID, 10+=filiere codes with priority rank values.

        Filiere columns: cell value = priority number (1=first choice, 2=second, …),
        None/empty = not chosen.
        """
        if not file_path.lower().endswith(".xlsx"):
            raise ValueError("Only .xlsx files are supported")
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        # Read header row to get filiere codes (columns 10+, 0-indexed: 9+)
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        filiere_headers = []  # list of (col_index, filiere_code) for cols 9+
        if header_row:
            for i, h in enumerate(header_row):
                if i >= 9 and h is not None:
                    filiere_headers.append((i, str(h).strip()))

        result = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[1]:
                continue  # skip empty Nom

            # Extract filiere choices: collect (priority, code), sort by priority asc
            choices = []
            for col_idx, fcode in filiere_headers:
                if col_idx < len(row) and row[col_idx] is not None:
                    try:
                        rank = int(row[col_idx])
                        if rank > 0:
                            choices.append((rank, fcode))
                    except (ValueError, TypeError):
                        pass
            choices.sort(key=lambda x: x[0])  # sort by priority rank (1 first)

            r = ExcelService.ImportRow(
                nom=str(row[1] or "").strip(),
                nom_intermediaire=str(row[2] or "").strip() if len(row) > 2 else "",
                prenom=str(row[3] or "").strip() if len(row) > 3 else "",
                date_de_naissance=str(row[4] or "").strip() if len(row) > 4 else "",
                sexe=str(row[5] or "").strip() if len(row) > 5 else "",
                langue=str(row[6] or "").strip() if len(row) > 6 else "fr",
                statut=str(row[7] or "").strip() if len(row) > 7 else "I",
                etab_code=str(row[8] or "").strip() if len(row) > 8 else "",
                filiere_choices=[(fcode, rank) for rank, fcode in choices],
            )
            result.append(r)
        wb.close()
        return result
